#!/usr/bin/env python3
"""
build_dataset.py — Cesvi Indice Infanzia
Estrae dati da 'Tabelle indicatori.xlsx' e produce:
  data/cesvi-indiceinfanzia_long.csv  (territory, year, indicator, value, rank)
"""

import re
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE = Path(__file__).parent
TABLES_FILE = BASE / "data/original/tables/Tabelle indicatori.xlsx"
OUTPUT_FILE = BASE / "data/cesvi-indiceinfanzia_long.csv"

# ── Region normalisation ───────────────────────────────────────────────────────
REGION_MAP = {
    "TRENTINO ALTO ADIGE":    "Trentino-Alto Adige",
    "TRENTINO-ALTO ADIGE":    "Trentino-Alto Adige",
    "FRIULI VENEZIA GIULIA":  "Friuli-Venezia Giulia",
    "FRIULI-VENEZIA GIULIA":  "Friuli-Venezia Giulia",
    "VALLE D'AOSTA":          "Valle d'Aosta",
    "EMILIA ROMAGNA":         "Emilia-Romagna",
    "EMILIA-ROMAGNA":         "Emilia-Romagna",
    "ABRUZZO":    "Abruzzo",    "BASILICATA": "Basilicata",
    "CALABRIA":   "Calabria",   "CAMPANIA":   "Campania",
    "LAZIO":      "Lazio",      "LIGURIA":    "Liguria",
    "LOMBARDIA":  "Lombardia",  "MARCHE":     "Marche",
    "MOLISE":     "Molise",     "PIEMONTE":   "Piemonte",
    "PUGLIA":     "Puglia",     "SARDEGNA":   "Sardegna",
    "SICILIA":    "Sicilia",    "TOSCANA":    "Toscana",
    "UMBRIA":     "Umbria",     "VENETO":     "Veneto",
}
ALL_REGIONS = set(REGION_MAP.values())


def norm_region(s):
    if not s:
        return None
    return REGION_MAP.get(str(s).strip().upper(), None)


# ── Sub-table label sequence (15 per year, in sheet order) ────────────────────
SUB_LABELS = [
    "cap_cura_rischio",               # Tab  1: Fattori rischio — Cura di sé
    "cap_cura_serv_maltrattanti",     # Tab  2: Servizi maltrattanti — Cura
    "cap_cura_serv_infanzia",         # Tab  3: Servizi infanzia — Ricevere cura
    "cap_vita_sana_rischio",          # Tab  4: Fattori rischio — Vita sana
    "cap_vita_sana_sintomi",          # Tab  5: Sintomi vulnerabilità — Vita sana
    "cap_vita_sana_serv_maltrattanti",# Tab  6: Servizi maltrattanti — Vita sana
    "cap_vita_sana_serv_infanzia",    # Tab  7: Servizi infanzia — Vita sana
    "cap_vita_sicura_rischio",        # Tab  8: Fattori rischio — Vita sicura
    "cap_vita_sicura_servizi",        # Tab  9: Servizi — Vita sicura
    "cap_conoscenza_rischio",         # Tab 10: Fattori rischio — Conoscenza e sapere
    "cap_conoscenza_servizi",         # Tab 11: Servizi — Conoscenza e sapere
    "cap_lavorare_rischio",           # Tab 12: Fattori rischio — Lavorare
    "cap_lavorare_servizi",           # Tab 13: Servizi — Lavorare
    "cap_accedere_risorse_rischio",   # Tab 14: Fattori rischio — Accedere alle risorse
    "cap_accedere_risorse_servizi",   # Tab 15: Servizi — Accedere alle risorse
]

# ── Aggregate definitions ──────────────────────────────────────────────────────
# capacity = average of its sub-component z-scores
CAP_COMPONENTS = {
    "cap_cura":               ["cap_cura_rischio",
                               "cap_cura_serv_maltrattanti",
                               "cap_cura_serv_infanzia"],
    "cap_vita_sana":          ["cap_vita_sana_rischio",
                               "cap_vita_sana_sintomi",
                               "cap_vita_sana_serv_maltrattanti",
                               "cap_vita_sana_serv_infanzia"],
    "cap_vita_sicura":        ["cap_vita_sicura_rischio",
                               "cap_vita_sicura_servizi"],
    "cap_conoscenza_sapere":  ["cap_conoscenza_rischio",
                               "cap_conoscenza_servizi"],
    "cap_lavorare":           ["cap_lavorare_rischio",
                               "cap_lavorare_servizi"],
    "cap_accedere_risorse":   ["cap_accedere_risorse_rischio",
                               "cap_accedere_risorse_servizi"],
}

# rischio = average of rischio sub-components per capacity (one per cap)
RISCHIO_CAPS = [
    "cap_cura_rischio",
    # vita_sana rischio = avg of tab4 + tab5 → use the pre-computed cap_vita_sana as proxy
    "cap_vita_sana_rischio",
    "cap_vita_sana_sintomi",
    "cap_vita_sicura_rischio",
    "cap_conoscenza_rischio",
    "cap_lavorare_rischio",
    "cap_accedere_risorse_rischio",
]

PREVENZIONE_CAPS = [
    "cap_cura_serv_maltrattanti",
    "cap_cura_serv_infanzia",
    "cap_vita_sana_serv_maltrattanti",
    "cap_vita_sana_serv_infanzia",
    "cap_vita_sicura_servizi",
    "cap_conoscenza_servizi",
    "cap_lavorare_servizi",
    "cap_accedere_risorse_servizi",
]


# ── Extraction helpers ─────────────────────────────────────────────────────────

def extract_region_data(rows, section_idx):
    """
    Estrae (territory → (value, rank)) da una sezione INDICE REGIONALE.
    Cerca la riga 'Regioni' e itera le 20-25 righe dati seguenti.
    """
    # Find 'Regioni' marker row
    data_start = None
    for i in range(section_idx + 1, min(section_idx + 12, len(rows))):
        raw = str(rows[i][1] or "").strip().lower()
        if raw == "regioni":
            data_start = i + 1
            break
    if data_start is None:
        data_start = section_idx + 3   # fallback

    result = {}
    for i in range(data_start, min(data_start + 30, len(rows))):
        row = rows[i]
        raw_name = str(row[1] or "").strip()
        territory = norm_region(raw_name)
        if not territory:
            continue

        # Collect all numeric values in the row (cols C onward)
        nums = []
        for cell in row[2:]:
            if cell is None:
                continue
            try:
                nums.append(float(str(cell).replace(" ", "")))
            except (ValueError, TypeError):
                pass

        if len(nums) >= 2:
            value = nums[-2]   # penultimate = z-score total
            rank  = int(round(nums[-1]))   # last = positional rank
            result[territory] = (value, rank)

    return result


def is_totale_section(rows, idx):
    """True se la sezione è un sommario TOTALE CAPACITÀ (non una delle 15 sub-tabelle)."""
    v2 = str(rows[idx][2] or "").strip().upper()
    return "TOTALE CAPACITÀ" in v2 or "TOTALE CAPACITY" in v2


def find_indice_sections(rows):
    """Restituisce gli indici di riga delle sezioni INDICE REGIONALE."""
    hits = []
    for i, row in enumerate(rows):
        v1 = str(row[1] or "").strip().upper()
        if "INDICE REGIONALE" in v1:
            hits.append(i)
    return hits


def process_sheet(ws, year):
    rows = list(ws.iter_rows(values_only=True))
    sections = find_indice_sections(rows)

    totale_secs = [i for i in sections if is_totale_section(rows, i)]
    normal_secs = [i for i in sections if not is_totale_section(rows, i)]

    print(f"  {year}: {len(normal_secs)} sub-tabelle normali, {len(totale_secs)} TOTALE")

    records = []

    # ── 15 normal sub-tables ──────────────────────────────────────────────────
    for tab_i, sec_idx in enumerate(normal_secs[:15]):
        label = SUB_LABELS[tab_i]
        data  = extract_region_data(rows, sec_idx)
        for territory, (value, rank) in data.items():
            records.append(dict(territory=territory, year=year,
                                indicator=label, value=value, rank=rank))

    # ── TOTALE sections (2018: rischio + prevenzione aggregate) ──────────────
    for sec_idx in totale_secs:
        v1 = str(rows[sec_idx][1] or "").upper()
        if "FATTORI" in v1:
            label = "indice_rischio"
        else:
            label = "indice_prevenzione"
        data = extract_region_data(rows, sec_idx)
        for territory, (value, rank) in data.items():
            records.append(dict(territory=territory, year=year,
                                indicator=label, value=value, rank=rank))

    return records


# ── Main ───────────────────────────────────────────────────────────────────────

def avg_cols(df, cols):
    available = [c for c in cols if c in df.columns]
    if not available:
        return float("nan")
    return df[available].mean(axis=1)


def add_ranks(df, col, by="year"):
    """Aggiunge colonna rank_<col> calcolata descending per year."""
    df[f"rank_{col}"] = df.groupby(by)[col].rank(ascending=False, method="min").astype("Int64")
    return df


if __name__ == "__main__":
    wb = openpyxl.load_workbook(str(TABLES_FILE), read_only=True, data_only=True)

    all_records = []
    for sheet_name in wb.sheetnames:
        year = int(sheet_name)
        ws   = wb[sheet_name]
        recs = process_sheet(ws, year)
        all_records.extend(recs)

    df_raw = pd.DataFrame(all_records)
    print(f"\nRecord grezzi estratti: {len(df_raw)}")

    # ── Pivot to wide for aggregate computation ────────────────────────────────
    dfp = df_raw.pivot_table(
        index=["territory", "year"],
        columns="indicator",
        values="value",
        aggfunc="first",
    ).copy()

    # ── 6 capacity aggregates ──────────────────────────────────────────────────
    for cap, comps in CAP_COMPONENTS.items():
        dfp[cap] = avg_cols(dfp, comps)

    # ── rischio/prevenzione sub-index ─────────────────────────────────────────
    # Use pre-extracted values for 2018; compute for other years
    dfp["indice_rischio_calc"]    = avg_cols(dfp, RISCHIO_CAPS)
    dfp["indice_prevenzione_calc"] = avg_cols(dfp, PREVENZIONE_CAPS)

    if "indice_rischio" in dfp.columns:
        # Fill gaps (non-2018 years) with computed values
        dfp["indice_rischio"] = dfp["indice_rischio"].combine_first(dfp["indice_rischio_calc"])
    else:
        dfp["indice_rischio"] = dfp["indice_rischio_calc"]

    if "indice_prevenzione" in dfp.columns:
        dfp["indice_prevenzione"] = dfp["indice_prevenzione"].combine_first(dfp["indice_prevenzione_calc"])
    else:
        dfp["indice_prevenzione"] = dfp["indice_prevenzione_calc"]

    # ── indice_totale = average of the two subindices ─────────────────────────
    dfp["indice_totale"] = avg_cols(dfp, ["indice_rischio", "indice_prevenzione"])

    # Drop helper columns
    dfp.drop(columns=["indice_rischio_calc", "indice_prevenzione_calc"], inplace=True, errors="ignore")

    # ── Melt back to long format ───────────────────────────────────────────────
    SUMMARY_INDICATORS = [
        "indice_totale", "indice_rischio", "indice_prevenzione",
        "cap_cura", "cap_vita_sana", "cap_vita_sicura",
        "cap_conoscenza_sapere", "cap_lavorare", "cap_accedere_risorse",
    ]
    ALL_INDICATORS = SUMMARY_INDICATORS + SUB_LABELS
    keep = [c for c in ALL_INDICATORS if c in dfp.columns]

    df_long = (
        dfp[keep]
        .reset_index()
        .melt(id_vars=["territory", "year"], var_name="indicator", value_name="value")
        .dropna(subset=["value"])
    )

    # ── Compute ranks ─────────────────────────────────────────────────────────
    # For raw sub-labels: use original ranks from extraction
    raw_ranks = (
        df_raw[df_raw["indicator"].isin(SUB_LABELS)]
        [["territory", "year", "indicator", "rank"]]
    )

    # For summary indicators: recompute rank by descending value within year
    summary_ranks = (
        df_long[df_long["indicator"].isin(SUMMARY_INDICATORS)]
        .assign(rank=lambda d: d.groupby(["year", "indicator"])["value"]
                .rank(ascending=False, method="min").astype("Int64"))
        [["territory", "year", "indicator", "rank"]]
    )

    all_ranks = pd.concat([raw_ranks, summary_ranks], ignore_index=True)
    df_out = df_long.merge(all_ranks, on=["territory", "year", "indicator"], how="left")
    df_out["value"] = df_out["value"].round(4)
    df_out["rank"]  = df_out["rank"].astype("Int64")
    df_out = df_out.sort_values(["year", "indicator", "rank"]).reset_index(drop=True)

    # ── Verify vs known 2018 values ────────────────────────────────────────────
    print("\n--- Verifica 2018 indice_rischio (atteso: Trentino=1.011) ---")
    chk = df_out[(df_out.year == 2018) & (df_out.indicator == "indice_rischio")].sort_values("rank")
    print(chk[["territory", "value", "rank"]].head(5).to_string(index=False))

    print("\n--- Verifica 2018 indice_prevenzione (atteso: Emilia-Romagna=1.236) ---")
    chk2 = df_out[(df_out.year == 2018) & (df_out.indicator == "indice_prevenzione")].sort_values("rank")
    print(chk2[["territory", "value", "rank"]].head(5).to_string(index=False))

    # ── Summary ────────────────────────────────────────────────────────────────
    pivot_check = df_out.groupby(["year", "indicator"]).size().unstack(fill_value=0)
    print(f"\nRecord finali: {len(df_out)}")
    print(pivot_check.to_string())

    df_out.to_csv(str(OUTPUT_FILE), index=False)
    print(f"\n✓ Salvato: {OUTPUT_FILE}")
