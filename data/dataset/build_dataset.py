"""
Estrae i dati dalle tabelle Excel dell'Indice Infanzia Cesvi
e produce data.csv e metadata.csv.
Adattato per supportare i file copiati da Word (ricerca flessibile delle colonne).
"""

import openpyxl
import pandas as pd
from pathlib import Path

# ── Percorsi ─────────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent
EXCEL = ROOT / "dati.xlsx"
OUT_DATA = ROOT / "cesvi-indiceinfanzia_data.csv"
OUT_META = ROOT / "cesvi-indiceinfanzia_metadata.csv"

YEARS = ["2018", "2019", "2020", "2021", "2022", "2024", "2026"]


# ── Mappature ─────────────────────────────────────────────────────────────────

def map_index(tabella_title: str) -> str:
    """Ricava il codice breve dell'indice dal titolo della TABELLA."""
    t = tabella_title.upper()
    if "SINTOMI" in t:
        return "rischio_bambini"
    # Controlla SERVIZI prima di FATTORI DI RISCHIO per evitare falsi positivi
    # (es. 2018 "INDICE REGIONALE SERVIZI FATTORI DI RISCHIO POTENZIALI MALTRATTANTI")
    if "SERVIZI PER BAMBINI" in t or "SERVIZI PER L'INFANZIA" in t or "SERVIZI PER L\u2019INFANZIA" in t or "SERVIZI BAMBINI" in t:
        return "servizi_bambini"
    if "SERVIZI PER POTENZIALI" in t or "SERVIZI POTENZIALI" in t or ("SERVIZI" in t and "POTENZIALI MALTRATTANTI" in t):
        return "servizi_adulti"
    if "FATTORI DI RISCHIO" in t:
        return "rischio_adulti"
    raise ValueError(f"Indice non riconosciuto: {tabella_title!r}")


def map_capacity(cap_name: str) -> str:
    """Ricava il codice breve della capacità dal nome completo."""
    c = cap_name.upper().replace('\n', ' ')
    if "RICEVERE CURA" in c:
        return "cura"
    if "CURA DI S" in c:          # CURA DI SÉ E DEGLI ALTRI
        return "cura"
    if "VITA SANA" in c:
        return "vita_sana"
    if "VITA SICURA" in c:
        return "vita_sicura"
    if "CONOSCENZA" in c:
        return "conoscenza_sapere"
    if "LAVORARE" in c:
        return "lavorare"
    if "ACCEDERE" in c:
        return "accedere_risorse"
    raise ValueError(f"Capacità non riconosciuta: {cap_name!r}")


def normalize_territory(territory: str) -> str:
    """Uniforma i nomi delle regioni con doppia denominazione."""
    t_upper = territory.upper()
    if "FRIULI" in t_upper:
        return "Friuli-Venezia Giulia"
    if "TRENTINO" in t_upper:
        return "Trentino-Alto Adige/Südtirol"
    if "AOSTA" in t_upper:
        return "Valle d'Aosta/Vallée d'Aoste"
    if "EMILIA" in t_upper:
        return "Emilia-Romagna"
    return territory


# ── Estrazione ────────────────────────────────────────────────────────────────

def _is_totale(v) -> bool:
    """Verifica se un valore di cella corrisponde a 'Totale', tollerando spazi
    interni dovuti al word-wrap di Excel (es. 'Total e')."""
    if not isinstance(v, str):
        return False
    normalized = ''.join(v.lower().split())
    return normalized in ('totale', 'tot')


def extract_year(ws, year: str) -> list[dict]:
    """Estrae tutti i record da un foglio (formato 2019-2024 e 2026).

    Strategia:
    1. Trova la riga TABELLA → index_code
    2. Cerca CAPACITÀ nel titolo stesso, poi nelle 3 righe successive → cap_code
    3. Trova la riga INDICE REGIONALE (entro 50 righe) → âncora per i dati
    4. Dall'âncora (inclusa), cerca "Totale" e "Regioni" entro 10 righe
    5. Estrae le 20 righe dati
    """
    rows = list(ws.iter_rows(values_only=True))
    records = []
    i = 0

    while i < len(rows):
        row = rows[i]

        # 1. Cerca TABELLA in qualsiasi colonna
        tabella_title = None
        for cell in row:
            if isinstance(cell, str) and cell.strip().upper().startswith("TABELLA"):
                tabella_title = cell.strip()
                break

        if not tabella_title:
            i += 1
            continue

        try:
            index_code = map_index(tabella_title)
        except ValueError:
            i += 1
            continue

        # 2. Cerca CAPACITÀ prima nel titolo stesso (2026: titolo include "CAPACITÀ DI...")
        # poi nelle 3 righe successive (formato storico)
        cap_code = None
        try:
            cap_code = map_capacity(tabella_title)
        except ValueError:
            pass
        if not cap_code:
            for offset in range(1, 4):
                if i + offset < len(rows):
                    for cell in rows[i + offset]:
                        if isinstance(cell, str):
                            try:
                                cap_code = map_capacity(cell.strip())
                                break
                            except ValueError:
                                pass
                if cap_code:
                    break

        if not cap_code:
            i += 1
            continue

        # 3. Cerca la riga INDICE REGIONALE entro 50 righe dalla TABELLA.
        # Il titolo della sezione dati è SEMPRE in colonna 0; le altre colonne
        # possono contenere "Indice regionale ..." come nome di indicatore
        # (false positive da ignorare).
        indice_row = None
        for j in range(i + 1, min(i + 50, len(rows))):
            cell0 = rows[j][0]
            if isinstance(cell0, str) and "INDICE REGIONALE" in cell0.strip().upper():
                indice_row = j
                break

        if indice_row is None:
            i += 1
            continue

        # 4. Dall'âncora (inclusa) cerca "Totale" e "Regioni" entro 20 righe.
        # Usa _is_totale() per tollerare word-wrap Excel (es. "Total e").
        # La finestra è 20 perché alcune tabelle (es. 2026 tab.6) hanno header
        # multi-riga molto lunghe ("Regioni" fino a 14 righe dopo l'ancora).
        totale_col = None
        regioni_row = None

        for j in range(indice_row, min(indice_row + 20, len(rows))):
            r = rows[j]
            if totale_col is None:
                for k, v in enumerate(r):
                    if _is_totale(v):
                        totale_col = k
                        break
            if r[0] and isinstance(r[0], str) and r[0].strip().upper() in ("REGIONI", "REGIONE"):
                regioni_row = j
                break

        if regioni_row is None or totale_col is None:
            i += 1
            continue

        # 4b. Sanity check: in alcuni fogli l'header "Totale" e "Posizione Tot."
        # sono scambiati rispetto ai dati (es. 2019 tab.6).
        # Se i primi valori a totale_col sono tutti interi in [1,20] → è la
        # colonna rank; lo score è una colonna più a destra.
        _sk = regioni_row + 1
        _sample: list = []
        while _sk < len(rows) and len(_sample) < 5:
            _dr = rows[_sk]; _sk += 1
            if not _dr[0] or not isinstance(_dr[0], str) or not _dr[0].strip():
                continue
            if totale_col < len(_dr):
                _sample.append(_dr[totale_col])
        if _sample and all(isinstance(v, int) and 1 <= v <= 20 for v in _sample):
            totale_col += 1

        # 5. Estrai le righe dati (territorio sempre in colonna 0)
        k = regioni_row + 1
        while k < len(rows):
            dr = rows[k]
            territory = dr[0] if len(dr) > 0 else None

            # Salta righe vuote
            if territory is None or not isinstance(territory, str) or not territory.strip():
                k += 1
                continue

            territory = territory.strip()

            # Condizioni di fine blocco
            if (territory.startswith("[")
                    or territory.upper().startswith("TABELLA")
                    or "INDICE REGIONALE" in territory.upper()
                    or territory.upper() in ("REGIONI", "REGIONE")):
                break

            score = dr[totale_col] if totale_col < len(dr) else None
            if score is not None:
                score_num = None
                if isinstance(score, (int, float)):
                    score_num = float(score)
                elif isinstance(score, str):
                    try:
                        score_num = float(score.strip().replace(',', '.'))
                    except ValueError:
                        pass
                if score_num is not None:
                    records.append({
                        "territory": normalize_territory(territory),
                        "year": int(year),
                        "capacity": cap_code,
                        "index": index_code,
                        "score": score_num,
                    })
            k += 1

        i = k

    return records


def extract_year_2018(ws) -> list[dict]:
    """Estrae i record dal foglio 2018, che ha un formato diverso dai successivi.

    Struttura 2018: ogni sezione inizia con una riga in cui
      col[0] = titolo dell'indice ("INDICE ...")  e
      col[1] = nome della capacità ("CAPACITÀ DI ...").
    Non esiste un header "TABELLA". Alcune sezioni aggregate ("TOTALE CAPACITÀ")
    vanno saltate. Il titolo dell'indice può essere spezzato su più righe
    consecutive (col[0] non-None, col[1]=None).
    """
    rows = list(ws.iter_rows(values_only=True))
    records = []
    i = 0

    while i < len(rows):
        row = rows[i]
        col0 = row[0] if len(row) > 0 else None
        col1 = row[1] if len(row) > 1 else None

        # Cerca righe-intestazione: col[0] ha "INDICE", col[1] ha "CAPACITÀ"
        if not (isinstance(col0, str) and "INDICE" in col0.upper() and
                isinstance(col1, str) and "CAPACITÀ" in col1.upper()):
            i += 1
            continue

        # Salta sezioni aggregate (TOTALE CAPACITÀ)
        if "TOTALE" in col1.upper() and "CAPACITÀ" in col1.upper():
            i += 1
            continue

        # Ricava la capacità da col[1]
        try:
            cap_code = map_capacity(col1.strip())
        except ValueError:
            i += 1
            continue

        # Ricostruisce il titolo completo dell'indice (può spezzarsi su più righe:
        # righe successive hanno col[0] non-None e col[1]=None)
        title = col0.strip()
        for offset in range(1, 4):
            if i + offset >= len(rows):
                break
            next_row = rows[i + offset]
            next_col0 = next_row[0] if len(next_row) > 0 else None
            next_col1 = next_row[1] if len(next_row) > 1 else None
            if isinstance(next_col0, str) and next_col0.strip() and next_col1 is None:
                title += " " + next_col0.strip()
            else:
                break

        try:
            index_code = map_index(title)
        except ValueError:
            i += 1
            continue

        # Trova la colonna "Totale" (o "TOT") nella stessa riga o nelle 3 successive
        totale_col = None
        for j_off in range(0, 4):
            if i + j_off >= len(rows):
                break
            for k, v in enumerate(rows[i + j_off]):
                if _is_totale(v):
                    totale_col = k
                    break
            if totale_col is not None:
                break

        if totale_col is None:
            i += 1
            continue

        # Trova la riga "Regioni" entro le 5 righe successive all'intestazione
        regioni_row = None
        for j in range(i + 1, min(i + 6, len(rows))):
            r = rows[j]
            if r[0] and isinstance(r[0], str) and r[0].strip().upper() in ("REGIONI", "REGIONE"):
                regioni_row = j
                break

        if regioni_row is None:
            i += 1
            continue

        # Estrae le righe dati
        k = regioni_row + 1
        while k < len(rows):
            dr = rows[k]
            territory = dr[0] if len(dr) > 0 else None

            if territory is None or not isinstance(territory, str) or not territory.strip():
                k += 1
                continue

            territory = territory.strip()

            if (territory.startswith("[")
                    or "INDICE" in territory.upper()
                    or territory.upper() in ("REGIONI", "REGIONE")):
                break

            score = dr[totale_col] if totale_col < len(dr) else None
            if score is not None:
                score_num = None
                if isinstance(score, (int, float)):
                    score_num = float(score)
                elif isinstance(score, str):
                    try:
                        score_num = float(score.strip().replace(',', '.'))
                    except ValueError:
                        pass
                if score_num is not None:
                    records.append({
                        "territory": normalize_territory(territory),
                        "year": 2018,
                        "capacity": cap_code,
                        "index": index_code,
                        "score": score_num,
                    })
            k += 1

        i = k

    return records


# ── Metadata ─────────────────────────────────────────────────────────────────

METADATA = [
    { "field": "rischio_adulti", "prefix": "Indice regionale", "label": "Fattori di Rischio Potenziali Maltrattanti" },
    { "field": "servizi_adulti", "prefix": "Indice regionale", "label": "Servizi per Potenziali Maltrattanti" },
    { "field": "servizi_bambini", "prefix": "Indice regionale", "label": "Servizi per Bambini/e" },
    { "field": "rischio_bambini", "prefix": "Indice regionale", "label": "Sintomi Potenziale Maltrattamento dei Bambini/e" },
    { "field": "cura", "prefix": "Capacità di", "label": "Cura" },
    { "field": "vita_sana", "prefix": "Capacità di", "label": "Vivere una Vita Sana" },
    { "field": "vita_sicura", "prefix": "Capacità di", "label": "Vivere una Vita Sicura" },
    { "field": "conoscenza_sapere", "prefix": "Capacità di", "label": "Acquisire Conoscenza e Sapere" },
    { "field": "lavorare", "prefix": "Capacità di", "label": "Lavorare" },
    { "field": "accedere_risorse", "prefix": "Capacità di", "label": "Accedere alle Risorse e ai Servizi" },
    # Indici aggregati
    { "field": "rischio", "prefix": "Indice regionale", "label": "Fattori di rischio" },
    { "field": "servizi", "prefix": "Indice regionale", "label": "Servizi" },
    { "field": "totale", "prefix": "Indice regionale", "label": "Maltrattamento e Cura all'Infanzia in Italia" },
]


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not EXCEL.exists():
        print(f"ERRORE: File '{EXCEL.name}' non trovato nella cartella {ROOT}")
        return

    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)

    # ── Patch manuale: record assenti nel file sorgente ──────────────────────
    # Sicilia 2018, servizi_adulti/conoscenza_sapere: dato mancante nella fonte
    MANUAL_RECORDS = [
        {"territory": "Sicilia", "year": 2018, "capacity": "conoscenza_sapere",
         "index": "servizi_adulti", "score": -1.756},
    ]

    all_records = []
    for year in YEARS:
        if year not in wb.sheetnames:
            print(f"Foglio '{year}' non trovato, salto.")
            continue
        ws = wb[year]
        if year == "2018":
            records = extract_year_2018(ws)
        else:
            records = extract_year(ws, year)
        print(f"{year}: {len(records)} record estratti")
        all_records.extend(records)

    all_records.extend(MANUAL_RECORDS)

    if not all_records:
        print("\nATTENZIONE: Nessun record estratto. Controlla il formato del file Excel.")
        return

    df = pd.DataFrame(all_records, columns=["territory", "year", "capacity", "index", "score"])

    # ── Divide la colonna 'index' in 'population' e 'type' ──────────────────
    # es. "rischio_adulti" → population="adulti", type="rischio"
    _INDEX_SPLIT = {
        "rischio_adulti":  ("adulti",  "rischio"),
        "rischio_bambini": ("bambini", "rischio"),
        "servizi_adulti":  ("adulti",  "servizi"),
        "servizi_bambini": ("bambini", "servizi"),
    }
    df["population"] = df["index"].map(lambda x: _INDEX_SPLIT[x][0])
    df["type"]       = df["index"].map(lambda x: _INDEX_SPLIT[x][1])
    df = df.drop(columns=["index"])
    df = df[["territory", "year", "population", "type", "capacity", "score"]]

    # ── Calcola i tre indici aggregati ───────────────────────────────────────
    # sentinel "totale" su population, type e capacity per le righe aggregate
    totale  = (df.groupby(["territory", "year", "capacity"])["score"].mean()
                 .groupby(["territory", "year"]).mean().rename("totale"))
    rischio = (df[df["type"] == "rischio"]
                 .groupby(["territory", "year"])["score"].mean().rename("rischio"))
    servizi = (df[df["type"] == "servizi"]
                 .groupby(["territory", "year"])["score"].mean().rename("servizi"))

    agg_rows = []
    for agg_type, series in [("totale", totale), ("rischio", rischio), ("servizi", servizi)]:
        for (territory, year), score in series.items():
            agg_rows.append({"territory": territory, "year": year,
                             "population": "totale", "type": agg_type,
                             "capacity": "totale", "score": score})

    df_agg = pd.DataFrame(agg_rows, columns=["territory", "year", "population", "type", "capacity", "score"])
    df = pd.concat([df, df_agg], ignore_index=True)

    # ── Arrotonda tutti gli score a 3 decimali ────────────────────────────────
    df["score"] = df["score"].round(3)

    df = df.sort_values(["year", "type", "population", "capacity", "territory"]).reset_index(drop=True)
    df.to_csv(OUT_DATA, index=False)
    print(f"\nSalvato: {OUT_DATA.name}  ({len(df)} righe)")

    df_meta = pd.DataFrame(METADATA, columns=["field", "prefix", "label"])
    df_meta.to_csv(OUT_META, index=False)
    print(f"Salvato: {OUT_META.name}  ({len(df_meta)} righe)")


if __name__ == "__main__":
    main()