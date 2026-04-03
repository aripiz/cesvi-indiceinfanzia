"""
Estrae i dati dalle tabelle Excel dell'Indice Infanzia Cesvi
e produce data.csv e metadata.csv.
Adattato per supportare i file copiati da Word (ricerca flessibile delle colonne).
"""

import openpyxl
import pandas as pd
from pathlib import Path

# ── Percorsi ─────────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent
# Aggiornato al nuovo nome del file che hai menzionato
EXCEL = ROOT / "original" / "dati.xlsx"
OUT_DATA = ROOT / "cesvi-indiceinfanzia_data.csv"
OUT_META = ROOT / "cesvi-indiceinfanzia_metadata.csv"

YEARS = ["2019", "2020", "2021", "2022", "2024"]


# ── Mappature ─────────────────────────────────────────────────────────────────

def map_index(tabella_title: str) -> str:
    """Ricava il codice breve dell'indice dal titolo della TABELLA."""
    t = tabella_title.upper()
    if "SINTOMI" in t:
        return "rischio_bambini"
    if "FATTORI DI RISCHIO" in t:
        return "rischio_adulti"
    if "SERVIZI PER BAMBINI" in t or "SERVIZI PER L'INFANZIA" in t or "SERVIZI PER L\u2019INFANZIA" in t:
        return "servizi_bambini"
    if "SERVIZI PER POTENZIALI" in t or "SERVIZI POTENZIALI" in t:
        return "servizi_adulti"
    raise ValueError(f"Indice non riconosciuto: {tabella_title!r}")


def map_capacity(cap_name: str) -> str:
    """Ricava il codice breve della capacità dal nome completo."""
    c = cap_name.upper().replace('\n', ' ')
    if "RICEVERE CURA" in c:
        return "cura"
    if "CURA DI S" in c:          # CURA DI SÉ E DEGLI ALTRI
        return "cura"
    if "VITA SANA" in c:
        return "vita_sana"
    if "VITA SICURA" in c:
        return "vita_sicura"
    if "CONOSCENZA" in c:
        return "conoscenza_sapere"
    if "LAVORARE" in c:
        return "lavorare"
    if "ACCEDERE" in c:
        return "accedere_risorse"
    raise ValueError(f"Capacità non riconosciuta: {cap_name!r}")


def normalize_territory(territory: str) -> str:
    """Uniforma i nomi delle regioni con doppia denominazione."""
    t_upper = territory.upper()
    if "FRIULI" in t_upper:
        return "Friuli-Venezia Giulia"
    if "TRENTINO" in t_upper:
        return "Trentino-Alto Adige/Südtirol"
    if "AOSTA" in t_upper:
        return "Valle d'Aosta/Vallée d'Aoste"
    return territory


# ── Estrazione ────────────────────────────────────────────────────────────────

def extract_year(ws, year: str) -> list[dict]:
    """Estrae tutti i record da un foglio.

    Strategia:
    1. Trova la riga TABELLA → index_code
    2. Trova CAPACITÀ entro le 3 righe successive → cap_code
    3. Trova la riga INDICE REGIONALE (entro 50 righe) → âncora per i dati
    4. Dall'âncora (inclusa), cerca "Totale" e "Regioni" entro 10 righe
    5. Estrae le 20 righe dati
    """
    rows = list(ws.iter_rows(values_only=True))
    records = []
    i = 0

    while i < len(rows):
        row = rows[i]

        # 1. Cerca TABELLA in qualsiasi colonna
        tabella_title = None
        for cell in row:
            if isinstance(cell, str) and cell.strip().upper().startswith("TABELLA"):
                tabella_title = cell.strip()
                break

        if not tabella_title:
            i += 1
            continue

        try:
            index_code = map_index(tabella_title)
        except ValueError:
            i += 1
            continue

        # 2. Cerca CAPACITÀ nelle 3 righe successive
        cap_code = None
        for offset in range(1, 4):
            if i + offset < len(rows):
                for cell in rows[i + offset]:
                    if isinstance(cell, str):
                        try:
                            cap_code = map_capacity(cell.strip())
                            break
                        except ValueError:
                            pass
            if cap_code:
                break

        if not cap_code:
            i += 1
            continue

        # 3. Cerca la riga INDICE REGIONALE entro 50 righe dalla TABELLA.
        # Il titolo della sezione dati è SEMPRE in colonna 0; le altre colonne
        # possono contenere "Indice regionale ..." come nome di indicatore
        # (false positive da ignorare).
        indice_row = None
        for j in range(i + 1, min(i + 50, len(rows))):
            cell0 = rows[j][0]
            if isinstance(cell0, str) and "INDICE REGIONALE" in cell0.strip().upper():
                indice_row = j
                break

        if indice_row is None:
            i += 1
            continue

        # 4. Dall'âncora (inclusa) cerca "Totale" e "Regioni" entro 10 righe
        # "Totale" può stare nella stessa riga dell'INDICE REGIONALE oppure in
        # una delle righe di intestazione immediatamente successive.
        totale_col = None
        regioni_row = None

        for j in range(indice_row, min(indice_row + 10, len(rows))):
            r = rows[j]
            if totale_col is None:
                for k, v in enumerate(r):
                    if isinstance(v, str) and v.strip() == "Totale":
                        totale_col = k
                        break
            if r[0] and isinstance(r[0], str) and r[0].strip().upper() in ("REGIONI", "REGIONE"):
                regioni_row = j
                break

        if regioni_row is None or totale_col is None:
            i += 1
            continue

        # 5. Estrai le righe dati (territorio sempre in colonna 0)
        k = regioni_row + 1
        while k < len(rows):
            dr = rows[k]
            territory = dr[0] if len(dr) > 0 else None

            # Salta righe vuote
            if territory is None or not isinstance(territory, str) or not territory.strip():
                k += 1
                continue

            territory = territory.strip()

            # Condizioni di fine blocco
            if (territory.startswith("[")
                    or territory.upper().startswith("TABELLA")
                    or "INDICE REGIONALE" in territory.upper()
                    or territory.upper() in ("REGIONI", "REGIONE")):
                break

            score = dr[totale_col] if totale_col < len(dr) else None
            if score is not None:
                score_num = None
                if isinstance(score, (int, float)):
                    score_num = float(score)
                elif isinstance(score, str):
                    try:
                        score_num = float(score.strip().replace(',', '.'))
                    except ValueError:
                        pass
                if score_num is not None:
                    records.append({
                        "territory": normalize_territory(territory),
                        "year": int(year),
                        "capacity": cap_code,
                        "index": index_code,
                        "score": score_num,
                    })
            k += 1

        i = k

    return records


# ── Metadata ─────────────────────────────────────────────────────────────────

METADATA = [
    { "field": "rischio_adulti", "prefix": "Indice regionale", "label": "Fattori di Rischio Potenziali Maltrattanti" },
    { "field": "servizi_adulti", "prefix": "Indice regionale", "label": "Servizi per Potenziali Maltrattanti" },
    { "field": "servizi_bambini", "prefix": "Indice regionale", "label": "Servizi per Bambini/e" },
    { "field": "rischio_bambini", "prefix": "Indice regionale", "label": "Sintomi Potenziale Maltrattamento dei Bambini/e" },
    { "field": "cura", "prefix": "Capacità di", "label": "Cura" },
    { "field": "vita_sana", "prefix": "Capacità di", "label": "Vivere una Vita Sana" },
    { "field": "vita_sicura", "prefix": "Capacità di", "label": "Vivere una Vita Sicura" },
    { "field": "conoscenza_sapere", "prefix": "Capacità di", "label": "Acquisire Conoscenza e Sapere" },
    { "field": "lavorare", "prefix": "Capacità di", "label": "Lavorare" },
    { "field": "accedere_risorse", "prefix": "Capacità di", "label": "Accedere alle Risorse e ai Servizi" },
    # Indici aggregati
    { "field": "rischio", "prefix": "Indice regionale", "label": "Fattori di rischio" },
    { "field": "servizi", "prefix": "Indice regionale", "label": "Servizi" },
    { "field": "totale", "prefix": "Indice regionale", "label": "Maltrattamento e Cura all'Infanzia in Italia" },
]


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not EXCEL.exists():
        print(f"ERRORE: File '{EXCEL.name}' non trovato nella cartella {ROOT}")
        return

    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)

    all_records = []
    for year in YEARS:
        if year not in wb.sheetnames:
            print(f"Foglio '{year}' non trovato, salto.")
            continue
        ws = wb[year]
        records = extract_year(ws, year)
        print(f"{year}: {len(records)} record estratti")
        all_records.extend(records)

    if not all_records:
        print("\nATTENZIONE: Nessun record estratto. Controlla il formato del file Excel.")
        return

    df = pd.DataFrame(all_records, columns=["territory", "year", "capacity", "index", "score"])

    # ── Calcola i tre indici aggregati ───────────────────────────────────────
    totale  = (df.groupby(["territory", "year", "capacity"])["score"].mean()
                 .groupby(["territory", "year"]).mean().rename("totale"))
    rischio = (df[df["index"].str.startswith("rischio")]
                 .groupby(["territory", "year"])["score"].mean().rename("rischio"))
    servizi = (df[df["index"].str.startswith("servizi")]
                 .groupby(["territory", "year"])["score"].mean().rename("servizi"))

    agg_rows = []
    for agg_name, series in [("totale", totale), ("rischio", rischio), ("servizi", servizi)]:
        for (territory, year), score in series.items():
            agg_rows.append({"territory": territory, "year": year,
                             "capacity": "", "index": agg_name, "score": score})

    df_agg = pd.DataFrame(agg_rows, columns=["territory", "year", "capacity", "index", "score"])
    df = pd.concat([df, df_agg], ignore_index=True)

    # ── Arrotonda tutti gli score a 3 decimali ────────────────────────────────
    df["score"] = df["score"].round(3)

    df = df.sort_values(["year", "capacity", "index", "territory"]).reset_index(drop=True)
    df.to_csv(OUT_DATA, index=False)
    print(f"\nSalvato: {OUT_DATA.name}  ({len(df)} righe)")

    df_meta = pd.DataFrame(METADATA, columns=["field", "prefix", "label"])
    df_meta.to_csv(OUT_META, index=False)
    print(f"Salvato: {OUT_META.name}  ({len(df_meta)} righe)")


if __name__ == "__main__":
    main()